import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

df = pd.read_csv('voters_2026.csv')

# Add Status field based on data completeness
def get_status(row):
    issues = []
    
    # Check if this is an OCR-unreadable placeholder
    if row['Name'] == '[OCR not readable]':
        return 'OCR failed - manual entry needed'
    
    # Check each field
    if pd.isna(row['Serial No']) or row['Serial No'] == '':
        issues.append('Serial No missing')
    if pd.isna(row['Voter ID']) or str(row['Voter ID']).strip() == '':
        issues.append('Voter ID missing')
    if pd.isna(row['Name']) or str(row['Name']).strip() == '':
        issues.append('Name missing')
    if pd.isna(row['Relation Type']) or str(row['Relation Type']).strip() == '':
        issues.append('Relation Type missing')
    if pd.isna(row['Relation Name']) or str(row['Relation Name']).strip() == '':
        issues.append('Relation Name missing')
    if pd.isna(row['House No']) or str(row['House No']).strip() == '':
        issues.append('House No missing')
    if pd.isna(row['Age']) or str(row['Age']).strip() == '':
        issues.append('Age missing')
    if pd.isna(row['Sex']) or str(row['Sex']).strip() == '':
        issues.append('Sex missing')
    
    if not issues:
        return 'All correct'
    else:
        return ', '.join(issues)

df['Status'] = df.apply(get_status, axis=1)

# Count statuses
status_counts = df['Status'].value_counts()
print('Status distribution:')
for status, count in status_counts.items():
    print(f'  {status}: {count}')
print()
all_correct = (df['Status'] == 'All correct').sum()
with_issues = (df['Status'] != 'All correct').sum()
print(f'All correct: {all_correct}')
print(f'With issues: {with_issues}')

# Save CSV
df.to_csv('voters_2026.csv', index=False)
print('\nCSV updated')

# Save Excel with formatting
df.to_excel('voters_2026.xlsx', index=False, sheet_name='Voters')

# Format Excel
wb = load_workbook('voters_2026.xlsx')
ws = wb.active

# Header style
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Color code by status
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

status_col = df.columns.get_loc('Status') + 1  # 1-indexed

for row in range(2, ws.max_row + 1):
    status_cell = ws.cell(row=row, column=status_col)
    status = status_cell.value
    
    if status == 'All correct':
        status_cell.fill = green_fill
    elif status == 'OCR failed - manual entry needed':
        # Highlight entire row yellow
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = yellow_fill
    else:
        # Has some issues - light red for status cell
        status_cell.fill = red_fill

# Auto-fit columns
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

wb.save('voters_2026.xlsx')
print('Excel updated with Status field and color coding!')
