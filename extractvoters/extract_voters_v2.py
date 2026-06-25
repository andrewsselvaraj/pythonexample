"""
Extract voter data from Tamil electoral roll PDF using OCR.

Fields: Serial No, Parliament Constituency, Assembly Constituency,
        Part No, Section, Voter ID, Name, Relation Type, Relation Name,
        House No, Age, Sex
"""

import os
import re
import csv
import fitz          # pymupdf
import pytesseract
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Use the sample PDF in the workspace when available, but keep support for
# existing external paths if the user points to a different roll file.
DEFAULT_PDF_FILE = os.path.join(
    SCRIPT_DIR,
    '2026-EROLLGEN-S22-224-SIR-DraftRoll-Revision1-TAM-1-WI.pdf'
)
if not os.path.exists(DEFAULT_PDF_FILE):
    DEFAULT_PDF_FILE = r'd:\andrew\election\voterlist\2026-EROLLGEN-S22-229-SIR-FinalRoll-Revision1-TAM-7-WI.pdf'

PDF_FILE = os.environ.get('PDF_FILE', DEFAULT_PDF_FILE)
CSV_FILE = os.environ.get('CSV_FILE', os.path.join(SCRIPT_DIR, 'voters_2026.csv'))
EXCEL_FILE = os.environ.get('EXCEL_FILE', os.path.join(SCRIPT_DIR, 'voters_2026.xlsx'))

# Tesseract data may be installed in a different location on a developer PC.
TESSDATA_DIR = r'd:\andrew\election\tessdata'
if os.path.exists(TESSDATA_DIR):
    os.environ['TESSDATA_PREFIX'] = TESSDATA_DIR


def infer_constituencies_from_path(pdf_path, default_parliament='22-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø', default_assembly='229-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø'):
    """Infer parliament/assembly constituency from the PDF filename."""
    if not pdf_path:
        return default_parliament, default_assembly

    file_name = os.path.basename(pdf_path)
    match = re.search(r'S(\d+)-(\d+)', file_name, re.I)
    if match:
        parliament_num = match.group(1)
        assembly_num = match.group(2)
        return f'{parliament_num}-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø', f'{assembly_num}-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø'

    return default_parliament, default_assembly

DPI = 200   # 200 DPI works best for Tamil OCR

# ‚îÄ‚îÄ Voter block start patterns ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
# Handles OCR noise variants on first line of each voter entry:
#   "32 2210978360 | |"  (trailing OCR noise)
#   "254 2211829837."    (trailing dot)
#   "765 2 2ZK3705530"   (serial gender voter_id ‚Äì page-31 form)
#   "34 HFX1175306"      (normal)
#   "211"                (serial only, voter ID on next line)
RE_BLOCK_START = re.compile(
    r'^[#\s]*(\d{1,4})\s+(\d)\s+([A-Z0-9]{9,14})[\s|.]*$|'    # alt-1: serial gender voter_id
    r'^[#\s]*(\d{1,4})\s+([A-Z0-9]{8,14})[\s|.]*$|'           # alt-2: serial voter_id
    r'^[#\s]*(\d{1,4})[\s.]*$'                                 # alt-3: serial only
)

RE_VOTERID  = re.compile(r'\b([A-Z]{3}[0-9]{6,8})\b')
RE_VOTERID2 = re.compile(r'\b([A-Z0-9]{9,14})\b')   # fallback for OCR-corrupted IDs

# Voter field patterns (ZWJ ‚Äå may appear after consonants in OCR output)
RE_NAME     = re.compile(r'ýÆ™ýØÜýÆØýÆ∞ýØç[‚Äå\s]*:[‚Äå\s]*(.*)')
RE_FATHER   = re.compile(r'ýÆ§ýÆ®ýØçýÆ§ýØà(?:ýÆØýÆøýÆ©ýØç)?[‚Äå\s]*ýÆ™ýØÜýÆØýÆ∞ýØç[‚Äå\s]*:[‚Äå\s]*(.*)')
RE_HUSBAND  = re.compile(r'ýÆïýÆ£ýÆµýÆ∞ýØç[‚Äå\s]*ýÆ™ýØÜýÆØýÆ∞ýØç[‚Äå\s]*:[‚Äå\s]*(.*)')
RE_MOTHER   = re.compile(r'ýÆ§ýÆæýÆØýÆøýÆ©ýØç[‚Äå\s]*ýÆ™ýØÜýÆØýÆ∞ýØç[‚Äå\s]*:[‚Äå\s]*(.*)')
RE_HOUSE    = re.compile(r'ýÆµýØÄýÆüýØçýÆüýØÅ[‚Äå\s]*ýÆéýÆ£ýØç[‚Äå\s]*[:\./][‚Äå\s]*(\S+)')
RE_AGE      = re.compile(r'ýÆµýÆØýÆ§ýØÅ[‚Äå\s]*[:\./][‚Äå\s]*(\d+)')
RE_SEX      = re.compile(r'ýÆ™ýÆæýÆ≤ýÆøýÆ©ýÆÆýØç[‚Äå\s]*[:\./][‚Äå\s]*(ýÆÜýÆ£ýØç|ýÆ™ýØÜýÆ£ýØç|ýÆÆýØÇýÆ©ýØçýÆ±ýÆæýÆÆýØç)', re.UNICODE)

# Page header patterns
RE_PARTNO = re.compile(r'ýÆ™ýÆæýÆïýÆÆýØç[‚Äå\s]*ýÆéýÆ£ýØç[‚Äå\s]*[:\./]?[‚Äå\s]*(\d+)', re.UNICODE)
RE_SECTION = re.compile(r'ýÆ™ýÆøýÆ∞ýÆøýÆµýØÅ[‚Äå\s]*ýÆéýÆ£ýØç[‚Äå\s]*ýÆÆýÆ±ýØçýÆ±ýØÅýÆÆýØç[‚Äå\s]*ýÆ™ýØÜýÆØýÆ∞ýØç[‚Äå\s]*(.+)', re.UNICODE)

# Parliament constituency: ýÆ®ýÆæýÆüýÆæýÆ≥ýØÅýÆÆýÆ©ýØçýÆ±ýÆ§ýØç ýÆ§ýØäýÆïýØÅýÆ§ýÆøýÆØýÆøýÆ©ýØç ýÆéýÆ£ýØç ýÆÆýÆ±ýØçýÆ±ýØÅýÆÆýØç ýÆ™ýØÜýÆØýÆ∞ýØç : 22-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø
RE_PARLIAMENT = re.compile(
    r'ýÆ®ýÆæýÆüýÆæýÆ≥ýØÅýÆÆýÆ©ýØçýÆ±(?:ýÆ§ýØç)?[‚Äå\s]*(?:ýÆ§ýØäýÆïýØÅýÆ§ýÆøýÆØýÆøýÆ©ýØç|ýÆ§ýØäýÆïýØÅýÆ§ýÆø)[^\n]*?(\d+[‚Äå\s]*[-‚Äì][^\n]*)',
    re.UNICODE
)
# Assembly constituency: ýÆöýÆüýØçýÆüýÆÆýÆ©ýØçýÆ±ýÆ§ýØç ýÆ§ýØäýÆïýØÅýÆ§ýÆøýÆØýÆøýÆ©ýØç ýÆéýÆ£ýØç ýÆÆýÆ±ýØçýÆ±ýØÅýÆÆýØç ýÆ™ýØÜýÆØýÆ∞ýØç : 229-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø
RE_ASSEMBLY = re.compile(
    r'ýÆöýÆüýØçýÆüýÆÆýÆ©ýØçýÆ±(?:ýÆ§ýØç)?[‚Äå\s]*(?:ýÆ§ýØäýÆïýØÅýÆ§ýÆøýÆØýÆøýÆ©ýØç|ýÆ§ýØäýÆïýØÅýÆ§ýÆø)[^\n]*?(\d+[‚Äå\s]*[-‚Äì][^\n]*)',
    re.UNICODE
)

SEX_MAP = {'ýÆÜýÆ£ýØç': 'Male', 'ýÆ™ýØÜýÆ£ýØç': 'Female', 'ýÆÆýØÇýÆ©ýØçýÆ±ýÆæýÆÆýØç': 'Third Gender'}

# Fallback constants derived from the PDF filename (for example,
# S22-224 means parliament seat 22 and assembly constituency 224).
# These are used whenever the OCR header cannot be parsed so every row always
# carries constituency information.
DEFAULT_PARLIAMENT, DEFAULT_ASSEMBLY = infer_constituencies_from_path(
    PDF_FILE,
    default_parliament='22-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø',
    default_assembly='229-ýÆïýÆ©ýØçýÆ©ýÆøýÆØýÆæýÆïýØÅýÆÆýÆ∞ýÆø'
)


# ‚îÄ‚îÄ PDF rendering ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def render_page(page, dpi=DPI):
    """Render a PDF page to a PIL Image."""
    scale = dpi / 72.0
    mat = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=mat)
    return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)


def ocr_column(img_crop):
    """Run Tesseract OCR on a cropped column image."""
    return pytesseract.image_to_string(
        img_crop, lang='tam+eng',
        config='--psm 4 --oem 1'  # psm 4 = single column of uniform text
    )


def split_image_into_columns(img, num_cols=3):
    """Crop a page image into num_cols equal vertical strips."""
    w, h = img.size
    col_w = w // num_cols
    cols = []
    for i in range(num_cols):
        x0 = i * col_w
        x1 = (i + 1) * col_w if i < num_cols - 1 else w
        cols.append(img.crop((x0, 0, x1, h)))
    return cols


# ‚îÄ‚îÄ Parsing helpers ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def normalize_header_value(value):
    """Normalize OCR header values so they are readable and Tableau-friendly."""
    if not value:
        return ''
    cleaned = re.sub(r'\s+', ' ', value).strip()
    cleaned = cleaned.replace('‚Äå', '').replace(',', '-')
    cleaned = re.sub(r'^[\s:./\-]+', '', cleaned)
    cleaned = re.sub(r'[\s:./\-]+$', '', cleaned)
    return cleaned


def extract_header_value(text, regex):
    """Extract a value from the page header using a regex applied line-by-line."""
    for line in text.splitlines():
        cleaned = re.sub(r'\s+', ' ', line).strip()
        if not cleaned:
            continue
        m = regex.search(cleaned)
        if m:
            return normalize_header_value(m.group(1))
    m = regex.search(text)
    if m:
        return normalize_header_value(m.group(1))
    return ''


def parse_page_header(text):
    """
    Extract part no, section name, parliament constituency, and assembly
    constituency from the combined OCR text of all columns on a page.
    Returns (part_no, section, parliament, assembly) ‚Äì any may be ''.
    """
    part_no    = ''
    section    = ''
    parliament = ''
    assembly   = ''

    part_no = extract_header_value(text, RE_PARTNO)
    section = normalize_header_value(extract_header_value(text, RE_SECTION))
    parliament = extract_header_value(text, RE_PARLIAMENT)
    assembly = extract_header_value(text, RE_ASSEMBLY)

    if not parliament:
        parliament = infer_constituencies_from_path(PDF_FILE)[0]
    if not assembly:
        assembly = infer_constituencies_from_path(PDF_FILE)[1]

    return part_no, section, parliament, assembly


def split_column_into_voter_blocks(col_text):
    """
    Split a single column's OCR text into individual voter blocks.
    A new block starts whenever a line matches RE_BLOCK_START.
    """
    lines = col_text.split('\n')
    blocks = []
    current = []
    for line in lines:
        if not line.strip():
            continue
        if RE_BLOCK_START.match(line.strip()):
            if current:
                blocks.append('\n'.join(current))
            current = [line]
        else:
            current.append(line)
    if current:
        blocks.append('\n'.join(current))
    return blocks


def parse_voter_block(text):
    """
    Parse one voter block (OCR text) into a dict of fields.
    Returns None if no serial number can be found.
    """
    serial   = ''
    voter_id = ''

    first_line = text.split('\n')[0].strip()
    m = RE_BLOCK_START.match(first_line)
    if m:
        if m.group(1):      # alt-1: serial + gender digit + voter_id
            serial   = m.group(1)
            voter_id = m.group(3) or ''
        elif m.group(4):    # alt-2: serial + voter_id
            serial   = m.group(4)
            voter_id = m.group(5) or ''
        elif m.group(6):    # alt-3: serial only
            serial   = m.group(6)

    # Voter ID not on the first line ‚Äì search the whole block
    if not voter_id:
        vid = RE_VOTERID.search(text)
        if vid:
            voter_id = vid.group(1)
        else:
            vid2 = RE_VOTERID2.search(text)
            if vid2:
                voter_id = vid2.group(1)

    if not serial:
        return None

    # Name
    name = ''
    m = RE_NAME.search(text)
    if m:
        name = m.group(1).strip().replace('‚Äå', '').strip()

    # Relation (father / husband / mother ‚Äì first match wins)
    relation_type, relation_name = '', ''
    for pat, rtype in [(RE_FATHER, 'Father'), (RE_HUSBAND, 'Husband'), (RE_MOTHER, 'Mother')]:
        m = pat.search(text)
        if m:
            relation_type = rtype
            relation_name = m.group(1).strip().replace('‚Äå', '').strip()
            break

    # House No
    house = ''
    m = RE_HOUSE.search(text)
    if m:
        house = m.group(1).strip()

    # Age
    age = ''
    m = RE_AGE.search(text)
    if m:
        age = m.group(1)

    # Sex
    sex = ''
    m = RE_SEX.search(text)
    if m:
        sex = SEX_MAP.get(m.group(1).strip().replace('‚Äå', ''), m.group(1))

    return {
        'Serial No':     serial,
        'Voter ID':      voter_id,
        'Name':          name,
        'Relation Type': relation_type,
        'Relation Name': relation_name,
        'House No':      house,
        'Age':           age,
        'Sex':           sex,
    }


# ‚îÄ‚îÄ Main ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def main():
    doc = fitz.open(PDF_FILE)
    total_pages = doc.page_count
    print(f"PDF: {total_pages} pages")

    all_voters = []

    # Running state ‚Äì updated whenever a page header reveals new values;
    # seeded from filename-derived defaults so no row is ever left blank.
    current_parliament, current_assembly = infer_constituencies_from_path(
        PDF_FILE,
        default_parliament=DEFAULT_PARLIAMENT,
        default_assembly=DEFAULT_ASSEMBLY,
    )
    current_part       = ''
    current_section    = ''

    for page_num in range(3, total_pages):
        print(f"Processing page {page_num + 1}/{total_pages}...", end=' ', flush=True)
        page = doc[page_num]
        img  = render_page(page)

        cols      = split_image_into_columns(img, num_cols=3)
        col_texts = [ocr_column(col_img) for col_img in cols]

        # Parse header from the full page text before touching voter rows
        part_no, section, parliament, assembly = parse_page_header('\n'.join(col_texts))
        if part_no:
            current_part = part_no
        if section:
            current_section = section
        if parliament:
            current_parliament = parliament
        if assembly:
            current_assembly = assembly

        page_count = 0
        for col_text in col_texts:
            for block in split_column_into_voter_blocks(col_text):
                voter = parse_voter_block(block)
                if voter:
                    voter['Parliament Constituency'] = current_parliament
                    voter['Assembly Constituency']   = current_assembly
                    voter['Part No']                 = current_part
                    voter['Section']                 = current_section
                    all_voters.append(voter)
                    page_count += 1

        print(f"{page_count} voters" if page_count else "(no voters)")

    print(f"\nTotal voters extracted: {len(all_voters)}")

    if not all_voters:
        print("No voters extracted. Check OCR output.")
        return

    # Column order for output
    fieldnames = [
        'Serial No',
        'Parliament Constituency',
        'Assembly Constituency',
        'Part No',
        'Section',
        'Voter ID',
        'Name',
        'Relation Type',
        'Relation Name',
        'House No',
        'Age',
        'Sex',
    ]

    os.makedirs(os.path.dirname(CSV_FILE), exist_ok=True)
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

    # ‚îÄ‚îÄ CSV ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    with open(CSV_FILE, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        writer.writerows(all_voters)
    print(f"CSV saved: {CSV_FILE}")

    # ‚îÄ‚îÄ Excel ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Voters"

    center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left     = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin     = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill(start_color="EBF3FF", end_color="EBF3FF", fill_type="solid")

    # Column widths (one entry per field)
    col_widths = {
        'Serial No':               8,
        'Parliament Constituency': 28,
        'Assembly Constituency':   28,
        'Part No':                 8,
        'Section':                 40,
        'Voter ID':                14,
        'Name':                    28,
        'Relation Type':           14,
        'Relation Name':           28,
        'House No':                10,
        'Age':                      6,
        'Sex':                      8,
    }

    # Left-aligned columns (text-heavy)
    left_align_cols = {'Parliament Constituency', 'Assembly Constituency',
                       'Section', 'Name', 'Relation Name'}

    for col_i, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_i, value=field)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = col_widths.get(field, 12)
    ws.row_dimensions[1].height = 30

    for row_i, voter in enumerate(all_voters, start=2):
        fill = alt_fill if row_i % 2 == 0 else None
        for col_i, field in enumerate(fieldnames, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=voter.get(field, ''))
            cell.border    = thin
            cell.alignment = left if field in left_align_cols else center
            if fill:
                cell.fill = fill

    ws.freeze_panes = 'A3'
    wb.save(EXCEL_FILE)
    print(f"Excel saved: {EXCEL_FILE}")


if __name__ == '__main__':
    main()
