from openpyxl import load_workbook
wb = load_workbook(r'd:\andrew\python\pythonexample\extractvoters\2026-EROLLGEN-S22-224-SIR-DraftRoll-Revision1-TAM-1-WI.xlsx', data_only=True)
ws = wb.active
print(wb.sheetnames)
print(ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
    print(row)
