import tkinter as tk
from tkinter import filedialog, messagebox
import threading
import os
import sys
import re
from openpyxl import load_workbook, Workbook

# Import the main extraction logic from extract_voters_v2.py
import importlib.util
spec = importlib.util.spec_from_file_location("extract_voters_v2", os.path.join(os.path.dirname(__file__), "extract_voters_v2.py"))
extract_voters = importlib.util.module_from_spec(spec)
sys.modules["extract_voters_v2"] = extract_voters
spec.loader.exec_module(extract_voters)

class VoterExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Tamil Nadu Voter PDF Extractor")
        self.selected_files = []
        self.selected_excel_files = []

        self.frame = tk.Frame(root, padx=16, pady=16)
        self.frame.pack(fill=tk.BOTH, expand=True)

        self.select_btn = tk.Button(self.frame, text="Select PDF Files", command=self.select_files)
        self.select_btn.pack(fill=tk.X, pady=(0, 8))

        self.files_listbox = tk.Listbox(self.frame, selectmode=tk.EXTENDED, height=8)
        self.files_listbox.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        self.run_btn = tk.Button(self.frame, text="Extract Voters", command=self.run_extraction)
        self.run_btn.pack(fill=tk.X, pady=(0, 8))

        self.excel_label = tk.Label(self.frame, text="Excel files for combining", anchor="w")
        self.excel_label.pack(fill=tk.X, pady=(8, 4))

        self.browse_excel_btn = tk.Button(self.frame, text="Browse Excel", command=self.select_excel_files)
        self.browse_excel_btn.pack(fill=tk.X, pady=(0, 4))

        self.excel_text = tk.Text(self.frame, height=6, wrap=tk.WORD)
        self.excel_text.pack(fill=tk.BOTH, expand=True, pady=(0, 8))

        self.combine_excel_btn = tk.Button(self.frame, text="Combine Excel", command=self.combine_excel_files)
        self.combine_excel_btn.pack(fill=tk.X)

        self.status_label = tk.Label(self.frame, text="Ready", anchor="w")
        self.status_label.pack(fill=tk.X, pady=(8, 0))

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select Electoral Roll PDFs",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if files:
            self.selected_files = list(files)
            self.files_listbox.delete(0, tk.END)
            for f in self.selected_files:
                self.files_listbox.insert(tk.END, f)

    def run_extraction(self):
        if not self.selected_files:
            messagebox.showwarning("No Files", "Please select one or more PDF files.")
            return
        self.status_label.config(text="Extracting... Please wait.")
        self.run_btn.config(state=tk.DISABLED)
        threading.Thread(target=self._extract_worker, daemon=True).start()

    def select_excel_files(self):
        files = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls")]
        )
        if files:
            self.selected_excel_files = list(files)
            self.excel_text.delete("1.0", tk.END)
            self.excel_text.insert("1.0", "\n".join(self.selected_excel_files))

    def combine_excel_files(self):
        if not self.selected_excel_files:
            messagebox.showwarning("No Files", "Please select one or more Excel files first.")
            return

        self.status_label.config(text="Combining Excel files... Please wait.")
        self.combine_excel_btn.config(state=tk.DISABLED)
        try:
            output_dir = os.path.join(os.path.dirname(__file__), "combinedexcel")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, "combined_voters.xlsx")

            combined_headers = []
            combined_rows = []

            for excel_path in self.selected_excel_files:
                wb = load_workbook(excel_path, data_only=True, read_only=True)
                ws = wb.worksheets[0]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()

                header = []
                header_row = None
                for idx, row in enumerate(rows[:10]):
                    if not row:
                        continue
                    cleaned = ["" if cell is None else str(cell).strip() for cell in row]
                    if not any(cleaned):
                        continue
                    joined = " ".join(cleaned).lower()
                    if any(keyword in joined for keyword in [
                        "serial", "voter", "name", "part", "section", "parliament",
                        "assembly", "relation", "house", "age", "sex"
                    ]):
                        header = [cell for cell in cleaned if str(cell).strip()]
                        header_row = idx
                        break

                if not header:
                    header = [f"Column{i + 1}" for i in range(len(rows[0]) if rows else 0)]
                    header_row = 0

                for h in header:
                    if h not in combined_headers:
                        combined_headers.append(h)

                if header_row is None:
                    continue

                for row in rows[header_row + 1:]:
                    if not row:
                        continue
                    values = ["" if cell is None else str(cell) for cell in row]
                    if not any(values):
                        continue
                    if len(values) < len(header):
                        values += [""] * (len(header) - len(values))
                    row_map = {header[i]: values[i] for i in range(min(len(header), len(values)))}
                    combined_rows.append([row_map.get(col, "") for col in combined_headers])

            if not combined_headers:
                raise ValueError("No usable headers were found in the selected Excel files.")

            new_wb = Workbook()
            ws = new_wb.active
            ws.title = "Combined"
            ws.append(combined_headers)
            for row in combined_rows:
                if len(row) < len(combined_headers):
                    row += [""] * (len(combined_headers) - len(row))
                ws.append(row)

            for col in ws.columns:
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

            new_wb.save(output_path)
            messagebox.showinfo("Success", f"Combined Excel saved to:\n{output_path}")
            self.status_label.config(text="Combined Excel saved successfully.")
        except Exception as exc:
            messagebox.showerror("Combine Error", f"Unable to combine Excel files:\n{exc}")
            self.status_label.config(text="Combine failed.")
        finally:
            self.combine_excel_btn.config(state=tk.NORMAL)

    def _extract_worker(self):
        for pdf_path in self.selected_files:
            try:
                # Set the PDF_FILE and output paths for each file
                extract_voters.PDF_FILE = pdf_path
                base = os.path.splitext(os.path.basename(pdf_path))[0]
                # Try to extract constituency code and name from filename (e.g., 229-Kanniyakumari)
                import re
                m = re.search(r'(\d{2,4})[-_]?([\w\(\)\u0B80-\u0BFF]+)?', base)
                if m:
                    code = m.group(1)
                    name = m.group(2) or ''
                    # Clean up name: replace non-alphanum with underscore, strip underscores
                    name = re.sub(r'[^\w\u0B80-\u0BFF]+', '_', name).strip('_')
                    out_dir = os.path.join(os.path.dirname(pdf_path), f"output_{code}_{name}" if name else f"output_{code}")
                else:
                    out_dir = os.path.join(os.path.dirname(pdf_path), "output")
                os.makedirs(out_dir, exist_ok=True)
                extract_voters.CSV_FILE = os.path.join(out_dir, f"{base}_voters.csv")
                extract_voters.EXCEL_FILE = os.path.join(out_dir, f"{base}_voters.xlsx")
                extract_voters.main()
            except Exception as e:
                print(f"Error processing {pdf_path}: {e}")
        self.status_label.config(text="Done! Check output files in output_<constituency> folders.")
        self.run_btn.config(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = VoterExtractorGUI(root)
    root.mainloop()
